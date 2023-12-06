#Importing libraries 
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
import time 
import openpyxl
from openpyxl import Workbook

#Loading Selenium Webdriver 
driver= webdriver.Firefox()
wait = WebDriverWait(driver, 5)

#Opening Google maps 
driver.get("https://www.google.com/maps")
time.sleep(3)


#Finding the search box 
#driver.switch_to_default_content()
searchbox=driver.find_element(By.ID, 'searchboxinput')
location= "fruit and vegetable shops in malleshwaram"
searchbox.send_keys(location)
searchbox.send_keys(Keys.ENTER)
time.sleep(5)

entries=driver.find_elements(By.CLASS_NAME, 'Nv2PK')

# Scroll down to load more businesses
while True:
    # Scroll to the last entry
    driver.execute_script("arguments[0].scrollIntoView();", entries[-1])
    time.sleep(4)  # Give the page some time to load additional content

    # Update the entries list with the newly loaded elements
    new_entries = driver.find_elements(By.CLASS_NAME, 'Nv2PK') 

    # Check if no more entries are loaded
    if len(entries) == len(new_entries):
        break  # Break the loop if no new entries are loaded
    entries=new_entries


#Prepare the excel file using the Openpyxl  
try:
    wb= openpyxl.load_workbook("companies.xlsx")
    sheet = wb.create_sheet("companies")
except:
    wb = Workbook()
    wb.save("companies.xlsx")
    sheet = wb.create_sheet("companies")

sheet.append(["query","name","reviews","title","address","phone","link"])

#Extracting the information from the results  
for entry in entries:
    name= entry.find_element(By.CLASS_NAME, 'fontHeadlineSmall').text
    rev_address = entry.find_elements(By.CLASS_NAME, 'W4Efsd')
    reviews = rev_address[0].text
    title_address = rev_address[1].find_element(By.CLASS_NAME, 'W4Efsd').text.split(' · ')
    title =  title_address[0]
    address = title_address[-1]
    link = entry.find_element(By.XPATH, './/a').get_attribute('href')

    try:
        phone = entry.find_element(By.CLASS_NAME, 'UsdlK').text
    except:
        phone = "no contact number"
    
    try:
        sheet.append([location,name,reviews,title,address,phone,link])
    except IndexError:
        pass
 
#saving the excel file 
wb.save("companies.xlsx")